#!/usr/bin/env python3
# filepath: /Users/adam/Documents/Developer/MyGithub/Python_Study/Proficient/recursion.py
# coding: utf-8
speter = "-" * 10

# 第四篇
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from collections import Counter
from openpyxl.utils import get_column_letter
import math


# ========= 参数 =========

case_sensitive = False  # 是否区分大小写
numeric_decimal_places = 2  # 数值型 col_a 参与匹配时，保留的小数位数



def process_excel_data_with_keywords(input_file_name, col_a, col_b, dir="/Users/adam/Desktop/Processing/", type="xlsx", sheet_name = 0):
    """
    处理 Excel 文件中的数据，基于指定列的关键词进行匹配，并输出结果至新的 Excel 文件。

    本方法将对输入的 Excel 文件执行以下操作：
    1. 根据提供的文件名和类型（`type` 参数），读取位于指定目录（`dir`）的 Excel 文件。
    2. 从指定的两列（`col_a` 和 `col_b`）中提取数据。
    3. 根据第二列的关键词与第一列的数据进行匹配，找到包含关键词的行，并记录匹配信息。
    4. 输出处理结果到指定目录，并在文件名中加入 `_output` 后缀。
    5. 支持不同格式的文件类型（如 `.xlsx`）。

    参数:
    ----------
    input_file_name : str
        输入的 Excel 文件名（不包含扩展名），将根据此文件名从指定目录读取文件。
    
    col_a : str
        第一个需要匹配数据的列名。该列中的数据将与第二列中的关键词进行匹配。
    
    col_b : str
        第二个包含关键词的列名。方法将从此列提取所有关键词，并在第一列中进行匹配。

    dir : str, optional
        输入和输出文件的目录路径，默认为 "/Users/adam/Desktop/Processing/"。
    
    type : str, optional
        文件类型（扩展名），默认为 "xlsx"。支持其他文件类型（如 "csv"）时，只需修改此参数。
    
    sheet_name : int or str, optional
        Excel 表单的名称或索引，默认为 0，表示第一个工作表。如果是多个工作表的情况，可以提供表单名或索引。

    返回:
    -------
    None
        此方法无返回值。处理后的数据会被写入到新的 Excel 文件，文件名会在原文件名基础上加上 "_output" 后缀。

    异常:
    -------
    FileNotFoundError
        如果输入文件路径无效或文件不存在，将抛出此异常。
    
    ValueError
        如果列名无效或列不存在，将抛出此异常。

    示例:
    -------
    process_excel_data_with_keywords(
        input_file_name="欧盟查重",
        col_a="英文名称",
        col_b="Active substance",
        dir="/path/to/files/",
        type="xlsx",
        sheet_name=0
    )
    """
    input_file = f"{dir}{input_file_name}.{type}"
    output_file = f"{dir}{input_file_name}_output.{type}"
    
    # 1. 读取 Excel
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    # 获取 col_b 列的列索引（使用 df.columns 获取所有列名的位置）
    col_b_idx = df.columns.get_loc(col_b)
    # 将列索引转换为字母表示（如 0 -> A, 1 -> B, 2 -> C 等）
    col_b_letter = get_column_letter(col_b_idx + 1)  # openpyxl 是从 1 开始计算列号的
    # print(col_b_idx, col_b_letter, "idx")
    
    # 2. 文本预处理（统一小写 + 去空 + 处理 NaN）
    # 2.1 处理 col_a：逐行判断“是否为数值”
    # 设计说明：
    # - col_a 可能是“文本 + 数值”的混合列（真实业务常见）
    # - 每一行单独判断是否可转为数值
    #   - 能转：按 numeric_decimal_places 统一精度，用于数值匹配
    #   - 不能转：按文本处理，用于正则匹配
    def normalize_a_value(val):
        if pd.isna(val):
            return ""

        # 尝试作为数值处理
        try:
            num = float(val)
            # 数值型：采用“向下截断”而非四舍五入
            # 业务原因：
            # - 登记/成分/法规数据以“有效精度下限”为准
            # - 避免因进位导致本应等价的数据被判为不一致
            factor = 10 ** numeric_decimal_places
            truncated = math.floor(num * factor) / factor
            return f"{truncated:.{numeric_decimal_places}f}"
        except (ValueError, TypeError):
            # 非数值，按文本处理
            s = str(val).strip()
            return s.lower() if not case_sensitive else s

    a_series = df[col_a].apply(normalize_a_value)
    b_series = (
        df[col_b]
        .fillna("")
        .astype(str)
        .str.lower() if not case_sensitive else df[col_b].str.strip()
    )

    # 2.2 处理 col_b 的数值精度（向下截断）
    # 说明：
    # - 与 col_a 采用完全一致的截断规则
    # - 确保“同一精度语义下”的数值可比性
    def truncate_numeric(val):
        if pd.isna(val):
            return ""
        factor = 10 ** numeric_decimal_places
        truncated = math.floor(val * factor) / factor
        return f"{truncated:.{numeric_decimal_places}f}"

    b_numeric_series = pd.to_numeric(df[col_b], errors="coerce").apply(truncate_numeric)

    # 3. 判断“包含关系”（列对列：A 列 对 全部 B 列）
    # 说明：
    # - A 列逐行取值
    # - B 列作为“词库”，只要命中任意一个即算匹配
    # - 采用完整单词匹配，避免 rid 命中 imidacloprid

    # 构建英文词库（去空、去重），并记录来源单元格
    # 结构：keyword_map = { "imidacloprid": ["B2", "B15"], ... }
    keyword_map = {}

    for idx, b in enumerate(b_series, start=2):  # start=2 对应 Excel 行号
        if not b:
            continue
        for k in re.split(r"[;,/]", b):
            k = k.strip()
            if not k:
                continue
            # 使用传入的 col_b 来动态构建单元格（比如 col_b="Active substance" 时，记录 B2、B15）
            keyword_map.setdefault(k, []).append(f"{col_b_letter}{idx}")

    match_list = []
    hit_word_list = []
    empty_row_count = 0

    for row_idx, a in enumerate(a_series):
        if not a:
            match_list.append(False)
            hit_word_list.append("")
            empty_row_count += 1
            continue

        hit_items = []

        for k, cells in keyword_map.items():
            # —— 数值精度匹配（基于小数位数）——
            # 规则：
            # - a 与 b 都已按 numeric_decimal_places 统一格式
            # - 字符串完全一致，视为“数值等价”
            if a:
                for c in cells:
                    b_row = int(c[1:]) - 2
                    b_val = b_numeric_series.iloc[b_row]

                    if b_val and a == b_val:
                        hit_items.append(f"{a}({c})")

            # —— 原有文本正则匹配（兜底）——
            pattern = rf"\b{re.escape(k)}\b"
            if re.search(pattern, a):
                for c in cells:
                    hit_items.append(f"{k}({c})")

        if hit_items:
            match_list.append(True)
            hit_word_list.append("; ".join(sorted(set(hit_items))))
        else:
            match_list.append(False)
            hit_word_list.append("")

    df["match"] = match_list
    df["命中英文"] = hit_word_list

    # 统计信息
    total_rows = len(df)
    hit_count = sum(match_list)

    print("==== 统计结果 ====")
    print(f"总行数: {total_rows}")
    print(f"命中数: {hit_count}")
    print(f"空行数: {empty_row_count}")
    print(f"结果存储：{output_file}")
    print("=================")

    # 4. 写回 Excel
    df.drop(columns=["match"]).to_excel(output_file, index=False)

    # 5. 使用 openpyxl 高亮匹配单元格
    wb = load_workbook(output_file)
    ws = wb.active

    highlight = PatternFill(fill_type="solid", fgColor="FFF59D")  # 浅黄色

    for row_idx, is_match in enumerate(df["match"], start=2):
        if is_match:
            ws.cell(row=row_idx, column=1).fill = highlight  # 高亮 A 列

    wb.save(output_file)

if __name__ == "__main__":
    print(f"{speter*2}Starting{speter*2}")
    try:
        process_excel_data_with_keywords(input_file_name="欧盟查重", col_a="英文名称", col_b="Active substance")
    except KeyboardInterrupt:
        print(f"{speter*2}手动退出程序{speter*2}")
    finally:

        print(f"{speter*2}Finished{speter*2}")
